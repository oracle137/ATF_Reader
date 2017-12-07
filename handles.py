############################################################################### # _*_ coding: utf-8
import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill

# !Python27

minlist = []
maxlist = []

redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')
yellowFill = PatternFill(start_color='ffff00',
                         end_color='ffff00',
                         fill_type='solid')

orangeFill = PatternFill(start_color='ffa500',
                         end_color='ffa500',
                         fill_type='solid')

greenyellowFIll = PatternFill(start_color='adff2f',
                              end_color='adff2f',
                              fill_type='solid')

greenFill = PatternFill(start_color='008000',
                        end_color='008000',
                        fill_type='solid')


def find_col(col_name, ws):
    for c in range(1, ws.max_column+1):
        if ws.cell(row=1, column=c).value == col_name:
            return c
    return 0


def handle_program_lines(f_hpl, filename, name,savelocation):
    global minlist
    global maxlist
    minlist = []
    maxlist = []

    wb2bool = False
    wsbool = False
    ws0bool = False
    ws1bool = False
    ws2bool = False
    ws3bool = False
    ws4bool = False
    ws5bool = False
    ws6bool = False
    ws7bool = False
    ws8bool = False

    wb = Workbook()
    ws = wb.active
    ws.title = "Program Tests"
    ws0 = wb.create_sheet(title="Sample Setting")
    ws1 = wb.create_sheet(title="Lamp Settings")
    ws2 = wb.create_sheet(title="Point")
    ws3 = wb.create_sheet(title="ScanH")
    ws4 = wb.create_sheet(title="Zone")
    ws5 = wb.create_sheet(title="Retro1000")
    ws6 = wb.create_sheet(title="ScanV")
    ws7 = wb.create_sheet(title="Maximum")
    ws8 = wb.create_sheet(title="Color")

    wb2 = Workbook()
    wb2ws1 = wb2.active
    wb2ws1.title = "ScanX"

    word = ""
    while word != "[EndOfFile]\n":
        word = f_hpl.readline()
        if word == "[EndOfFile]\n":

            str1 = name + ".xlsx"
            str2 = name + "Scan" + ".xlsx"
            # wb.save(str1)
            if not wsbool:
                wb.remove_sheet(ws)
            if not ws0bool:
                wb.remove_sheet(ws0)
            if not ws1bool:
                wb.remove_sheet(ws1)
            if not ws2bool:
                wb.remove_sheet(ws2)
            if not ws3bool:
                wb.remove_sheet(ws3)
            if not ws4bool:
                wb.remove_sheet(ws4)
            if not ws5bool:
                wb.remove_sheet(ws5)
            if not ws6bool:
                wb.remove_sheet(ws6)
            if not ws7bool:
                wb.remove_sheet(ws7)
            if not ws8bool:
                wb.remove_sheet(ws8)

            if wb._sheets.__len__() > 0:
                wb.save(savelocation + "/" + str1)

            if wb2bool:
                wb2.save(savelocation + "/" + str2)
            # SINCE the atf reader can take in many different files it is setup to remove the old memory so data isnt over lapped
            sh = wb.get_sheet_names()
            for index in sh:
                a = index.encode('ascii', 'ignore')
                ws24 = wb.get_sheet_by_name(a)
                wb.remove_sheet(ws24)
            sh2 = wb2.get_sheet_names()
            for index in sh2:
                b = index.encode('ascii', 'ignore')
                ws24 = wb2.get_sheet_by_name(b)
                wb2.remove_sheet(ws24)
            return

        words = word.split()

        # if word.find("BeginSample") != -1:
        #     handle_sample(f_hpl, ws0)
        #     ws0bool = 1
        # elif word.find("BeginLamp") != -1:
        #     handle_lamp_tests(f_hpl, ws1)
        #     ws1bool = 1
        # elif word.find("[BeginProgram]") != -1:
        #     handle_programs(f_hpl, ws)
        #     wsbool = 1
        if len(words) > 1:
            if words[1] == "Point":
                handle(f_hpl, words[1], ws2)
                ws2bool = 1
            elif words[1] == "ScanH":
                handle(f_hpl, words[1], ws3)
                ws3bool = 1
            elif word.find("Zone=") != -1:
                handle(f_hpl, words[1], ws4)
                ws4bool = 1
            elif words[1] == "Retro1000":
                handle(f_hpl, words[1], ws5)
                ws5bool = 1
            elif words[1] == "ScanX":
                handle(f_hpl, words[1], wb2ws1)
            elif words[1] == "ScanV":
                handle(f_hpl, words[1], ws6)
                ws6bool = 1
            elif word == "PCode= Maximum\n":
                handle(f_hpl, words[1], ws7)
                ws7bool = 1
            elif words[1] == "Color":
                handle(f_hpl, words[1], ws8)
                ws8bool = 1
        elif word.find("H=") != -1:
            handle_the_scan(f_hpl, word, wb2)
            wb2bool = 1


def handle(f, name, ws):
    global minlist
    global maxlist
    lines = []
    ws.title = name
    word = ""

    while word != "{EndProgramLine}\n" and word != "{EndFunctionGroup}\n":
        word = f.readline()

        if word != "{EndProgramLine}\n" and word != "{EndFunctionGroup}\n":
            word = word.rstrip()
            word = word.rstrip('\xa1\xc6')
            word = word.rstrip('cd')
            word = word.rstrip('Maximum')
            word = word.rstrip('Minimum')
            lines.append(word)

    a = ws.max_row

    if a == 1:
        for i, line in enumerate(lines):
            line = line.split()
            line[0] = line[0].rstrip('=')
            ws.cell(row=1, column=i + 1).value = line[0]

    c = ws.max_row + 1

    for i, line in enumerate(lines):
        words = line.split()

        for word in words[1:]:
            col_num = find_col(words[0].rstrip('='), ws)
            if col_num == 0:
                col_num = ws.max_column + 1
                ws.cell(row=1,column=col_num).value = words[0].rstrip('=')
            if ws.cell(row=c, column=col_num).value is None:
                ws.cell(row=c, column=col_num).value = word.decode("utf-8", errors='ignore')
            else:
                ws.cell(row=c, column=col_num).value += " " + word.decode("utf-8", errors='ignore')

            if name == "ScanH" or name == "ScanX" or name == "Maximum":
                if words[0].rstrip('=') == "PMin":
                    minlist.append(int(word.decode("utf-8", errors='ignore')))
                if words[0].rstrip('=') == "PMax":
                    maxlist.append(int(word.decode("utf-8", errors='ignore')))


def handle_the_scan(f_hp, word_from_handle, wb2):
    global minlist
    global maxlist
    counter = 0
    # LOOPS used for extracting the data from the atf
    h = []
    v = []
    values = []
    word = word_from_handle
    ws12 = wb2.create_sheet()

    if word.find("H=") != -1:
        while word.find("V=") == -1:
            word = f_hp.readline()
            if word.find("V=") == -1:
                word = word.rstrip('\r\n')
                # you must STRIP THE '\n' from the '-45\n' so its counted as a int
                h.append(word)
    if word.find("V=") != -1:
        while word.find("{EndAngles}") == -1:
            word = f_hp.readline()
            if word.find("{EndAngles}") == -1:
                word = word.rstrip('\r\n')
                v.append(word)
            else:
                word = f_hp.readline()
                word = f_hp.readline()
                break
    if word.find("ValuesInCd") != -1:
        while word.find("EndValuesInCd") == -1:
            word = f_hp.readline()
            word = word.rstrip('\r\n')
            values.append(word)
    index = 1
    rindex = 0
    aboolean = False
    for i in h:
        b = ws12.max_row
        if b == 1:
            ws12.cell(row=b, column=index + 1).value = i
            index = index + 1
        else:
            if not aboolean:
                b = ws12.max_row + 1
                aboolean = True
            ws12.cell(row=b, column=index + 1).value = i
            index = index + 1
    for j in v:
        a = ws12.max_row
        ws12.cell(row=a + 1, column=1).value = j
    a = ws12.max_row
    a2 = a
    b = ws12.max_column
    if a > len(v) + 2:
        rindex = (ws12.max_row + 1 - len(v))
    else:
        rindex = 2

    if not values:
        return
    for hor in range(0, len(h)):
        for ver in range(0, len(v)):
            ws12.cell(row=ver + 2, column=hor + 2).value = values[(len(v) * hor) + ver]
            if (minlist[counter] == "-" or minlist[counter] == "" or int(minlist[counter] == 0)) and \
                    (maxlist[counter] == "-" or maxlist[counter] == "" or int(maxlist[counter] == 0)):
                pass
            elif minlist[counter] == "-" or minlist[counter] == "":
                if (float(values[(len(v) * hor) + ver]) > (float(maxlist[counter]) / 1.2)) and (
                            float(values[(len(v) * hor) + ver]) < (float(maxlist[counter]) / 1.15)):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = greenFill
                elif (float(values[(len(v) * hor) + ver]) > (float(maxlist[counter]) / 1.15)) and (
                            float(values[(len(v) * hor) + ver]) < (float(maxlist[counter]) / 1.10)):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = greenyellowFIll
                elif (float(values[(len(v) * hor) + ver]) > (float(maxlist[counter]) / 1.10)) and (
                            float(values[(len(v) * hor) + ver]) < (float(maxlist[counter]) / 1.05)):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = yellowFill
                elif (float(values[(len(v) * hor) + ver]) > (float(maxlist[counter]) / 1.05)) and (
                            float(values[(len(v) * hor) + ver]) < (float(maxlist[counter]) / 1.00)):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = orangeFill
                elif float(values[(len(v) * hor) + ver]) > float(maxlist[counter]):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = redFill
            elif maxlist[counter] == "-" or maxlist[counter] == "":
                if (float(values[(len(v) * hor) + ver]) < (float(minlist[counter]) * 1.2)) and (
                        (float(values[(len(v) * hor) + ver]) > (float(minlist[counter]) * 1.5))):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = greenFill
                elif (float(values[(len(v) * hor) + ver]) < (float(minlist[counter]) * 1.15)) and (
                        (float(values[(len(v) * hor) + ver]) > (float(minlist[counter]) * 1.10))):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = greenyellowFIll
                elif (float(values[(len(v) * hor) + ver]) < (float(minlist[counter]) * 1.10)) and (
                        (float(values[(len(v) * hor) + ver]) > (float(minlist[counter]) * 1.05))):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = yellowFill
                elif (float(values[(len(v) * hor) + ver]) < (float(minlist[counter]) * 1.05)) and (
                        (float(values[(len(v) * hor) + ver]) > (float(minlist[counter]) * 1.00))):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = orangeFill
                elif float(values[(len(v) * hor) + ver]) < float(minlist[counter]):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = redFill
            else:
                if (((float(values[(len(v) * hor) + ver]) < (float(minlist[counter]) * 1.2)) and (
                            float(values[(len(v) * hor) + ver]) > (float(minlist[counter]) * 1.15))) or (
                            (float(values[(len(v) * hor) + ver]) > (float(maxlist[counter]) / 1.2)) and (
                                (float(values[(len(v) * hor) + ver]) < (float(maxlist[counter]) / 1.15))))):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = greenFill
                elif (((float(values[(len(v) * hor) + ver]) < (float(minlist[counter]) * 1.15)) and (
                            float(values[(len(v) * hor) + ver]) > (float(minlist[counter]) * 1.10))) or (
                            (float(values[(len(v) * hor) + ver]) > (float(maxlist[counter]) / 1.15)) and (
                                (float(values[(len(v) * hor) + ver]) < (float(maxlist[counter]) / 1.10))))):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = greenyellowFIll
                elif (((float(values[(len(v) * hor) + ver]) < (float(minlist[counter]) * 1.10)) and (
                            float(values[(len(v) * hor) + ver]) > (float(minlist[counter]) * 1.05))) or (
                            (float(values[(len(v) * hor) + ver]) > (float(maxlist[counter]) / 1.10)) and (
                                (float(values[(len(v) * hor) + ver]) < (float(maxlist[counter]) / 1.05))))):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = yellowFill
                elif (((float(values[(len(v) * hor) + ver]) < (float(minlist[counter]) * 1.05) and (
                            float(values[(len(v) * hor) + ver]) > (float(minlist[counter]) * 1.00))) or (
                            (float(values[(len(v) * hor) + ver]) > (float(maxlist[counter]) / 1.05)) and (
                                (float(values[(len(v) * hor) + ver]) < (float(maxlist[counter]) / 1.00)))))):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = orangeFill
                elif ((float(values[(len(v) * hor) + ver]) < float(minlist[counter])) or (
                        (float(values[(len(v) * hor) + ver]) > float(maxlist[counter])))):
                    ws12.cell(row=ver + 2, column=hor + 2).fill = redFill
    if len(minlist) > 1:
        minlist.pop(0)
        maxlist.pop(0)
    else:
        pass
